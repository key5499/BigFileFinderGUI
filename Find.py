import sys
import os
import ctypes
import psutil
import subprocess
from ctypes import wintypes
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                               QHBoxLayout, QTreeView, QTableView, QSplitter,
                               QPushButton, QComboBox, QLabel, QProgressBar, 
                               QMessageBox, QMenu, QAbstractItemView,
                               QFrame, QGridLayout, QHeaderView, QStyle,
                               QStyleFactory, QStyledItemDelegate, QCheckBox)
from PySide6.QtCore import (Qt, QThread, Signal, QModelIndex, QDir, 
                           QSortFilterProxyModel, QPoint, QTimer, QSize,
                           QItemSelectionModel, QAbstractTableModel)
from PySide6.QtGui import (QStandardItemModel, QStandardItem, QAction, 
                          QFont, QColor, QBrush, QIcon, QPalette, QFontMetrics,
                          QPainter)

class FolderSizeScanner(QThread):
    """快速扫描文件夹大小的线程"""
    progress = Signal(str, int, int)  # 当前扫描路径，当前数量，总数量估算
    finished = Signal(list)           # 扫描完成
    error = Signal(str)               # 错误信号
    
    def __init__(self, root_path, scan_files=False, scan_folders=True):
        super().__init__()
        self.root_path = root_path
        self._cancelled = False
        self.scan_files = scan_files
        self.scan_folders = scan_folders
        
    def cancel(self):
        self._cancelled = True
        
    def run(self):
        try:
            results = []
            items = []
            
            # 收集所有需要扫描的项目
            if self.scan_folders:
                try:
                    for root, dirs, files in os.walk(self.root_path):
                        if self._cancelled:
                            return
                        items.append((root, 'folder'))
                        # 限制最大扫描文件夹数量，避免内存问题
                        if len(items) > 10000:
                            break
                except Exception as e:
                    print(f"遍历文件夹出错: {e}")
                    items.append((self.root_path, 'folder'))
            
            if self.scan_files:
                try:
                    for root, dirs, files in os.walk(self.root_path):
                        if self._cancelled:
                            return
                        for file in files:
                            file_path = os.path.join(root, file)
                            items.append((file_path, 'file'))
                            # 限制最大扫描文件数量，避免内存问题
                            if len(items) > 50000:
                                break
                        if len(items) > 50000:
                            break
                except Exception as e:
                    print(f"遍历文件出错: {e}")
            
            total_items = len(items)
            
            # 扫描每个项目
            for i, (item_path, item_type) in enumerate(items):
                if self._cancelled:
                    return
                    
                try:
                    if item_type == 'folder':
                        # 扫描文件夹大小
                        item_size = self._get_folder_size(item_path)
                        result = {
                            'type': 'folder',
                            'path': item_path,
                            'name': os.path.basename(item_path) if item_path != self.root_path else os.path.splitdrive(item_path)[0] + '根目录',
                            'size': item_size,
                            'display_size': self._format_size(item_size),
                            'level': item_path.count(os.sep) - self.root_path.count(os.sep)
                        }
                    else:
                        # 扫描文件大小
                        item_size = os.path.getsize(item_path)
                        result = {
                            'type': 'file',
                            'path': item_path,
                            'name': os.path.basename(item_path),
                            'size': item_size,
                            'display_size': self._format_size(item_size),
                            'level': item_path.count(os.sep) - self.root_path.count(os.sep)
                        }
                    
                    results.append(result)
                    
                    # 更新进度
                    progress = int((i + 1) * 100 / total_items) if total_items > 0 else 0
                    self.progress.emit(item_path, i + 1, total_items)
                    
                except (PermissionError, OSError) as e:
                    continue
                except Exception as e:
                    print(f"扫描项目 {item_path} 出错: {e}")
                    continue
            
            # 按大小排序
            results.sort(key=lambda x: x['size'], reverse=True)
            
            if not self._cancelled:
                self.finished.emit(results)
                
        except Exception as e:
            self.error.emit(str(e))
    
    def _get_folder_size(self, folder_path):
        """获取文件夹大小"""
        total_size = 0
        
        try:
            # 使用os.walk方法递归计算所有文件大小
            for dirpath, dirnames, filenames in os.walk(folder_path):
                if self._cancelled:
                    return 0
                for filename in filenames:
                    try:
                        filepath = os.path.join(dirpath, filename)
                        total_size += os.path.getsize(filepath)
                    except (OSError, PermissionError):
                        continue
        except (PermissionError, OSError):
            try:
                # 尝试使用os.scandir递归计算
                for entry in os.scandir(folder_path):
                    if self._cancelled:
                        return 0
                        
                    try:
                        if entry.is_file():
                            total_size += entry.stat().st_size
                        elif entry.is_dir():
                            # 递归计算子目录大小
                            try:
                                total_size += self._get_folder_size(entry.path)
                            except:
                                pass
                    except (OSError, PermissionError):
                        continue
            except:
                return 0
            
        return total_size
    
    def _format_size(self, size_bytes):
        """格式化文件大小显示"""
        if size_bytes == 0:
            return "0 B"
        
        size_names = ("B", "KB", "MB", "GB", "TB")
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1
        
        if i == 0:
            return f"{int(size_bytes)} B"
        elif i == 1:
            return f"{size_bytes:.1f} KB"
        elif i == 2:
            return f"{size_bytes:.1f} MB"
        elif i == 3:
            return f"{size_bytes:.2f} GB"
        else:
            return f"{size_bytes:.2f} TB"

class ItemSizeModel(QAbstractTableModel):
    """自定义表格模型，用于显示文件和文件夹大小"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.items = []
        self.headers = ['序号', '名称', '类型', '路径', '大小', '百分比']
        
    def rowCount(self, parent=None):
        return len(self.items)
    
    def columnCount(self, parent=None):
        return len(self.headers)
    
    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or index.row() >= len(self.items):
            return None
            
        item = self.items[index.row()]
        
        if role == Qt.DisplayRole:
            if index.column() == 0:  # 序号
                return str(index.row() + 1)
            elif index.column() == 1:  # 名称
                return item.get('name', '')
            elif index.column() == 2:  # 类型
                return "文件夹" if item.get('type') == 'folder' else "文件"
            elif index.column() == 3:  # 路径
                return item.get('path', '')
            elif index.column() == 4:  # 大小
                return item.get('display_size', '')
            elif index.column() == 5:  # 百分比
                return self._calculate_percentage(index.row())
                
        elif role == Qt.ForegroundRole:
            size_gb = item.get('size', 0) / (1024**3)
            if size_gb > 10:  # 大于10GB
                return QColor('#FF6B6B')  # 红色
            elif size_gb > 1:  # 大于1GB
                return QColor('#FFA726')  # 橙色
            elif size_gb > 0.1:  # 大于100MB
                return QColor('#FFEE58')  # 黄色
            else:
                return QColor('#FFFFFF')
                
        elif role == Qt.ToolTipRole:
            return f"路径: {item.get('path', '')}\n大小: {item.get('display_size', '')}\n类型: {'文件夹' if item.get('type') == 'folder' else '文件'}"
            
        elif role == Qt.UserRole:  # 用于排序的原始大小数据
            return item.get('size', 0)
            
        elif role == Qt.FontRole and index.column() == 1:  # 文件夹名称加粗
            font = QFont()
            if item.get('type') == 'folder':
                font.setBold(True)
            return font
            
        return None
    
    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.headers[section]
        return None
    
    def set_items(self, items):
        self.beginResetModel()
        self.items = items
        self.endResetModel()
    
    def _calculate_percentage(self, row_index):
        """计算项目大小占总扫描大小的百分比"""
        if not self.items or row_index >= len(self.items):
            return "0%"
        
        total_size = sum(f.get('size', 0) for f in self.items)
        if total_size == 0:
            return "0%"
            
        item_size = self.items[row_index].get('size', 0)
        percentage = (item_size / total_size) * 100
        return f"{percentage:.1f}%"

class SizeBarDelegate(QStyledItemDelegate):
    """自定义委托，显示大小条形图"""
    def paint(self, painter, option, index):
        if index.column() == 4:  # 大小列（现在是第5列，索引为4）
            # 获取原始模型和索引
            proxy_model = index.model()
            source_model = proxy_model.sourceModel()
            source_index = proxy_model.mapToSource(index)
            
            if not source_index.isValid():
                super().paint(painter, option, index)
                return
            
            # 获取原始大小数据
            item_data = source_model.items[source_index.row()]
            size_bytes = item_data.get('size', 0)
            
            # 计算最大值用于比例
            max_size = max((f.get('size', 0) for f in source_model.items), default=1)
            
            # 绘制背景
            painter.save()
            painter.setRenderHint(QPainter.Antialiasing)
            
            # 绘制背景矩形
            bg_rect = option.rect.adjusted(2, 2, -2, -2)
            painter.fillRect(bg_rect, QColor('#424242'))
            
            # 计算条形图宽度
            if max_size > 0:
                percentage = size_bytes / max_size
                bar_width = int(percentage * (bg_rect.width() - 4))
                
                # 根据大小设置颜色
                size_gb = size_bytes / (1024**3)
                if size_gb > 10:
                    bar_color = QColor('#FF5252')
                elif size_gb > 1:
                    bar_color = QColor('#FF9800')
                elif size_gb > 0.1:
                    bar_color = QColor('#FFEB3B')
                else:
                    bar_color = QColor('#4CAF50')
                
                # 绘制条形图
                bar_rect = bg_rect.adjusted(2, 2, -(bg_rect.width() - bar_width), -2)
                painter.fillRect(bar_rect, bar_color)
                
                # 添加圆角效果
                painter.setPen(Qt.NoPen)
                painter.setBrush(bar_color)
                painter.drawRoundedRect(bar_rect, 3, 3)
            
            # 绘制文本
            display_text = item_data.get('display_size', '')
            painter.setPen(QColor('#FFFFFF'))
            painter.drawText(bg_rect, Qt.AlignCenter, display_text)
            
            painter.restore()
        else:
            super().paint(painter, option, index)

class DarkDiskSpaceAnalyzer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.scanner_thread = None
        self.current_scan_path = ""
        self.init_ui()
        self.load_disks()
        
    def init_ui(self):
        """初始化用户界面 - 夜晚模式"""
        self.setWindowTitle('磁盘空间分析工具 - 夜晚模式')
        self.setGeometry(100, 100, 1600, 900)
        
        # 设置深色主题样式
        self.set_dark_theme()
        
        # 创建中心部件
        central_widget = QWidget()
        central_widget.setObjectName("centralWidget")
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(8)
        main_layout.setContentsMargins(12, 12, 12, 12)
        
        # ========== 顶部控制面板 ==========
        control_frame = QFrame()
        control_frame.setObjectName("controlFrame")
        control_frame.setFixedHeight(70)
        
        control_layout = QGridLayout(control_frame)
        control_layout.setSpacing(10)
        control_layout.setContentsMargins(10, 10, 10, 10)
        
        # 刷新磁盘按钮（最左边）
        self.refresh_button = QPushButton("🔄 刷新磁盘")
        self.refresh_button.clicked.connect(self.refresh_disks)
        self.refresh_button.setFixedWidth(120)
        self.refresh_button.setObjectName("refreshButton")
        control_layout.addWidget(self.refresh_button, 0, 0)
        
        # 磁盘选择
        control_layout.addWidget(QLabel("💾 磁盘:"), 0, 1)
        
        self.disk_combo = QComboBox()
        self.disk_combo.setFixedWidth(120)
        
        self.disk_combo.setObjectName("diskCombo")
        control_layout.addWidget(self.disk_combo, 0, 2)
        
        # 扫描方式复选框
        control_layout.addWidget(QLabel("扫描方式:"), 0, 3)
        
        self.scan_files_checkbox = QCheckBox("文件")
        self.scan_files_checkbox.setObjectName("scanFilesCheckbox")
        control_layout.addWidget(self.scan_files_checkbox, 0, 4)
        
        self.scan_folders_checkbox = QCheckBox("文件夹")
        self.scan_folders_checkbox.setObjectName("scanFoldersCheckbox")
        self.scan_folders_checkbox.setChecked(True)  # 默认扫描文件夹
        control_layout.addWidget(self.scan_folders_checkbox, 0, 5)
        
        # 中间留空（拉伸）
        control_layout.setColumnStretch(6, 1)
        
        # 右边操作按钮
        self.scan_button = QPushButton("🔍 开始扫描")
        self.scan_button.clicked.connect(self.start_scan)
        self.scan_button.setFixedWidth(120)
        self.scan_button.setObjectName("scanButton")
        control_layout.addWidget(self.scan_button, 0, 7)
        
        self.stop_button = QPushButton("⏹️ 停止扫描")
        self.stop_button.clicked.connect(self.stop_scan)
        self.stop_button.setEnabled(False)
        self.stop_button.setFixedWidth(120)
        self.stop_button.setObjectName("stopButton")
        control_layout.addWidget(self.stop_button, 0, 8)
        
        self.export_button = QPushButton("💾 导出列表")
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setFixedWidth(120)
        self.export_button.setObjectName("exportButton")
        self.export_button.setEnabled(False)  # 初始禁用，扫描完成后启用
        control_layout.addWidget(self.export_button, 0, 9)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(15)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setObjectName("progressBar")
        control_layout.addWidget(self.progress_bar, 1, 0, 1, 10)
        
        main_layout.addWidget(control_frame)
        
        # ========== 主内容区域 ==========
        main_splitter = QSplitter(Qt.Horizontal)
        main_splitter.setHandleWidth(3)
        
        # 左侧：文件夹树
        left_widget = QWidget()
        left_widget.setObjectName("leftWidget")
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(0)
        
        tree_label = QLabel("📁 文件夹结构")
        tree_label.setObjectName("sectionLabel")
        tree_label.setFixedHeight(30)
        left_layout.addWidget(tree_label)
        
        self.tree_view = QTreeView()
        self.tree_view.setObjectName("treeView")
        self.tree_view.setHeaderHidden(True)
        self.tree_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree_view.customContextMenuRequested.connect(self.show_tree_context_menu)
        self.tree_view.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tree_view.setAnimated(True)
        self.tree_view.setIndentation(15)
        
        self.tree_model = QStandardItemModel()
        self.tree_view.setModel(self.tree_model)
        
        left_layout.addWidget(self.tree_view)
        
        # 右侧：文件夹大小列表
        right_widget = QWidget()
        right_widget.setObjectName("rightWidget")
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(0)
        
        list_label = QLabel("📊 文件夹大小排序")
        list_label.setObjectName("sectionLabel")
        list_label.setFixedHeight(30)
        right_layout.addWidget(list_label)
        
        # 表格视图
        self.table_view = QTableView()
        self.table_view.setObjectName("tableView")
        self.table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self.show_table_context_menu)
        self.table_view.doubleClicked.connect(self.open_folder_from_table)
        self.table_view.setSortingEnabled(True)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.setSelectionMode(QAbstractItemView.ExtendedSelection)  # 支持Ctrl和Shift多选
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectRows)  # 按行选择
        
        # 自定义模型
        self.table_model = ItemSizeModel()
        self.table_proxy = QSortFilterProxyModel()
        self.table_proxy.setSourceModel(self.table_model)
        self.table_proxy.setSortRole(Qt.UserRole)
        self.table_view.setModel(self.table_proxy)
        
        # 设置列宽
        self.table_view.setColumnWidth(0, 60)   # 序号
        self.table_view.setColumnWidth(1, 200)  # 名称
        self.table_view.setColumnWidth(2, 80)   # 类型
        self.table_view.setColumnWidth(3, 400)  # 路径
        self.table_view.setColumnWidth(4, 150)  # 大小
        self.table_view.setColumnWidth(5, 80)   # 百分比
        
        # 设置大小列的委托
        self.table_view.setItemDelegateForColumn(4, SizeBarDelegate(self.table_view))
        
        right_layout.addWidget(self.table_view)
        
        # 添加到分割器
        main_splitter.addWidget(left_widget)
        main_splitter.addWidget(right_widget)
        
        # 设置分割器比例
        main_splitter.setStretchFactor(0, 3)  # 左侧占3份
        main_splitter.setStretchFactor(1, 7)  # 右侧占7份
        
        main_layout.addWidget(main_splitter, 1)
        
        # ========== 底部状态栏 ==========
        self.statusBar().showMessage("就绪")
        
        # 连接信号
        self.disk_combo.currentIndexChanged.connect(self.on_disk_changed)
        self.tree_view.expanded.connect(self.on_tree_item_expanded)
        
    def set_dark_theme(self):
        """设置深色主题"""
        self.setStyleSheet("""
            QMainWindow {
                background-color: #121212;
            }
            #centralWidget {
                background-color: #121212;
            }
            QFrame#controlFrame {
                background-color: #1E1E1E;
                border-radius: 6px;
                border: 1px solid #333333;
            }
            QLabel#sectionLabel {
                color: #BB86FC;
                font-weight: bold;
                font-size: 14px;
                padding-left: 10px;
                background-color: #1E1E1E;
                border-bottom: 1px solid #333333;
            }
            QLabel#statusLabel {
                color: #03DAC6;
                font-weight: bold;
            }
            QComboBox#diskCombo {
                background-color: #2D2D2D;
                color: #E0E0E0;
                border: 1px solid #444444;
                border-radius: 4px;
                padding: 5px 8px;
                min-height: 15px;
                selection-background-color: #BB86FC;
            }
            QComboBox#diskCombo:hover {
                border: 1px solid #BB86FC;
            }
            QComboBox#diskCombo::drop-down {
                border: none;
            }
            QComboBox#diskCombo QAbstractItemView {
                background-color: #2D2D2D;
                color: #E0E0E0;
                selection-background-color: #BB86FC;
                border: 1px solid #444444;
            }
            QPushButton {
                background-color: #2D2D2D;
                color: #E0E0E0;
                border: 1px solid #444444;
                border-radius: 4px;
                padding: 5px 8px;
                font-weight: bold;
                min-height: 15px;
            }
            QPushButton:hover {
                background-color: #3D3D3D;
                border: 1px solid #BB86FC;
            }
            QPushButton:pressed {
                background-color: #BB86FC;
                color: #121212;
            }
            QPushButton:disabled {
                background-color: #1A1A1A;
                color: #666666;
                border: 1px solid #333333;
            }
            QPushButton#scanButton {
                background-color: #1976D2;
            }
            QPushButton#scanButton:hover {
                background-color: #2196F3;
            }
            QPushButton#stopButton {
                background-color: #D32F2F;
            }
            QPushButton#stopButton:hover {
                background-color: #F44336;
            }
            QPushButton#openButton {
                background-color: #388E3C;
            }
            QPushButton#openButton:hover {
                background-color: #4CAF50;
            }
            QPushButton#refreshButton {
                background-color: #7B1FA2;
            }
            QPushButton#refreshButton:hover {
                background-color: #9C27B0;
            }
            QTreeView#treeView, QTableView#tableView {
                background-color: #1E1E1E;
                color: #E0E0E0;
                border: 1px solid #333333;
                border-radius: 4px;
                alternate-background-color: #252525;
                selection-background-color: #BB86FC;
                selection-color: #121212;
                outline: none;
            }
            QTreeView#treeView::item, QTableView#tableView::item {
                padding: 5px;
            }
            QTreeView#treeView::item:hover, QTableView#tableView::item:hover {
                background-color: #2D2D2D;
            }
            QTreeView#treeView::item:selected, QTableView#tableView::item:selected {
                background-color: #BB86FC;
                color: #121212;
            }
            QHeaderView::section {
                background-color: #2D2D2D;
                color: #E0E0E0;
                padding: 8px;
                border: 1px solid #333333;
                font-weight: bold;
            }
            QProgressBar#progressBar {
                background-color: #2D2D2D;
                border: 1px solid #444444;
                border-radius: 4px;
                text-align: center;
                color: #E0E0E0;
            }
            QProgressBar#progressBar::chunk {
                background-color: #03DAC6;
                border-radius: 4px;
            }
            QMenu {
                background-color: #2D2D2D;
                color: #E0E0E0;
                border: 1px solid #444444;
            }
            QMenu::item {
                padding: 8px 30px 8px 20px;
            }
            QMenu::item:selected {
                background-color: #BB86FC;
                color: #121212;
            }
            QMenu::separator {
                height: 1px;
                background-color: #444444;
                margin: 5px 10px;
            }
        """)
        
    def load_disks(self):
        """加载可用磁盘"""
        self.disk_combo.clear()
        
        disks = []
        for part in psutil.disk_partitions():
            try:
                if os.name == 'nt':  # Windows
                    if 'cdrom' in part.opts or not part.mountpoint:
                        continue
                    
                    # 获取磁盘信息
                    usage = psutil.disk_usage(part.mountpoint)
                    free_gb = usage.free / (1024**3)
                    total_gb = usage.total / (1024**3)
                    
                    # 格式化显示
                    if 'fixed' in part.opts:
                        icon = "💾"
                    elif 'removable' in part.opts:
                        icon = "💿"
                    else:
                        icon = "📀"
                        
                    display_text = f"{icon} {part.mountpoint} ({free_gb:.1f}GB 可用 / {total_gb:.1f}GB)"
                    
                    disks.append({
                        'path': part.mountpoint,
                        'display': display_text,
                        'usage': usage
                    })
            except Exception as e:
                print(f"加载磁盘 {part.mountpoint} 出错: {e}")
                continue
        
        # 按磁盘路径排序
        disks.sort(key=lambda x: x['path'])
        
        for disk in disks:
            self.disk_combo.addItem(disk['display'], disk['path'])
        
        if disks:
            # 默认选择C盘
            c_drive_index = self.disk_combo.findData("C:\\")
            if c_drive_index >= 0:
                self.disk_combo.setCurrentIndex(c_drive_index)
            else:
                self.disk_combo.setCurrentIndex(0)
    
    def on_disk_changed(self, index):
        """磁盘选择变化"""
        if index >= 0:
            disk_path = self.disk_combo.itemData(index)
            if disk_path:
                self.load_disk_tree(disk_path)
    
    def load_disk_tree(self, disk_path):
        """加载磁盘树形结构"""
        self.tree_model.clear()
        
        if not os.path.exists(disk_path):
            QMessageBox.warning(self, "警告", f"磁盘路径不存在: {disk_path}")
            return
        
        # 添加磁盘根节点
        usage = psutil.disk_usage(disk_path)
        used_percent = (usage.used / usage.total) * 100 if usage.total > 0 else 0
        
        disk_text = f"💾 {disk_path} - 已用 {used_percent:.1f}% ({self._format_size(usage.used)} / {self._format_size(usage.total)})"
        
        disk_item = QStandardItem(disk_text)
        disk_item.setData(disk_path, Qt.UserRole)
        disk_item.setEditable(False)
        
        # 添加一级子文件夹（延迟加载）
        try:
            for entry in os.scandir(disk_path):
                if entry.is_dir() and not entry.name.startswith('$') and not entry.name.startswith('.'):
                    try:
                        # 检查是否有子文件夹
                        has_children = False
                        for _ in os.scandir(entry.path):
                            pass
                        has_children = True
                        
                        folder_item = QStandardItem(f"📁 {entry.name}")
                        folder_item.setData(entry.path, Qt.UserRole)
                        folder_item.setEditable(False)
                        
                        # 如果有子文件夹，添加占位符
                        if has_children:
                            placeholder = QStandardItem("...")
                            placeholder.setEditable(False)
                            folder_item.appendRow(placeholder)
                        
                        disk_item.appendRow(folder_item)
                    except (PermissionError, OSError):
                        continue
                    except StopIteration:
                        # 没有子文件夹
                        folder_item = QStandardItem(f"📁 {entry.name}")
                        folder_item.setData(entry.path, Qt.UserRole)
                        folder_item.setEditable(False)
                        disk_item.appendRow(folder_item)
        except (PermissionError, OSError) as e:
            print(f"加载磁盘 {disk_path} 的子文件夹出错: {e}")
        
        self.tree_model.appendRow(disk_item)
        self.tree_view.expand(disk_item.index())
    
    def on_tree_item_expanded(self, index):
        """树节点展开时加载子文件夹"""
        item = self.tree_model.itemFromIndex(index)
        
        # 如果有占位符，则加载子文件夹
        if item.rowCount() == 1:
            child = item.child(0)
            if child and child.text() == "...":
                item.removeRow(0)
                self.load_subfolders(item)
    
    def load_subfolders(self, parent_item):
        """加载子文件夹"""
        path = parent_item.data(Qt.UserRole)
        
        try:
            for entry in os.scandir(path):
                if entry.is_dir() and not entry.name.startswith('$') and not entry.name.startswith('.'):
                    try:
                        # 检查是否有子文件夹
                        has_children = False
                        try:
                            next(os.scandir(entry.path))
                            has_children = True
                        except StopIteration:
                            has_children = False
                        except:
                            has_children = True
                        
                        folder_item = QStandardItem(f"📁 {entry.name}")
                        folder_item.setData(entry.path, Qt.UserRole)
                        folder_item.setEditable(False)
                        
                        # 如果有子文件夹，添加占位符
                        if has_children:
                            placeholder = QStandardItem("...")
                            placeholder.setEditable(False)
                            folder_item.appendRow(placeholder)
                        
                        parent_item.appendRow(folder_item)
                    except (PermissionError, OSError):
                        continue
        except (PermissionError, OSError):
            pass
    
    def start_scan(self):
        """开始扫描"""
        current_index = self.tree_view.currentIndex()
        
        if current_index.isValid():
            item = self.tree_model.itemFromIndex(current_index)
            scan_path = item.data(Qt.UserRole)
        else:
            # 如果没有选中节点，使用当前选中的磁盘
            disk_index = self.disk_combo.currentIndex()
            if disk_index >= 0:
                scan_path = self.disk_combo.itemData(disk_index)
            else:
                QMessageBox.warning(self, "警告", "请先选择磁盘或文件夹")
                return
        
        if not scan_path or not os.path.exists(scan_path):
            QMessageBox.warning(self, "警告", "选择的路径不存在")
            return
        
        self.current_scan_path = scan_path
        
        # 禁用扫描按钮，启用停止按钮
        self.scan_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.export_button.setEnabled(False)
        self.statusBar().showMessage("🔄 正在扫描...")
        self.progress_bar.setValue(0)
        
        # 清空表格
        self.table_model.set_items([])
        
        # 获取扫描方式
        scan_files = self.scan_files_checkbox.isChecked()
        scan_folders = self.scan_folders_checkbox.isChecked()
        
        # 创建并启动扫描线程
        self.scanner_thread = FolderSizeScanner(scan_path, scan_files, scan_folders)
        self.scanner_thread.progress.connect(self.update_progress)
        self.scanner_thread.finished.connect(self.scan_finished)
        self.scanner_thread.error.connect(self.scan_error)
        self.scanner_thread.start()
    
    def update_progress(self, current_path, current, total):
        """更新进度"""
        folder_name = os.path.basename(current_path)
        progress = int(current * 100 / total) if total > 0 else 0
        self.progress_bar.setValue(progress)
        self.statusBar().showMessage(f"🔍 完成进度 {progress}% ({current}/{total}) 正在扫描: {folder_name}...")
    
    def scan_finished(self, results):
        """扫描完成"""
        # 恢复按钮状态
        self.scan_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.export_button.setEnabled(True)
        # 统计文件和文件夹数量
        folder_count = sum(1 for r in results if r['type'] == 'folder')
        file_count = sum(1 for r in results if r['type'] == 'file')
        
        self.progress_bar.setValue(100)
        status_msg = f"✅ 扫描完成，共 {len(results)} 个项目（{folder_count} 个文件夹，{file_count} 个文件）"
        self.statusBar().showMessage(status_msg)
        
        # 将结果设置到表格模型
        self.table_model.set_items(results)
        self.table_proxy.sort(4, Qt.DescendingOrder)  # 按大小列（第5列，索引4）排序
        
        # 显示统计信息
        if results:
            total_size = sum(r['size'] for r in results)
            largest = results[0]['display_size'] if results else "0 B"
            largest_name = results[0]['name'] if results else ""
            
            msg = f"📊 扫描完成！\n\n"
            msg += f"📁 扫描路径: {self.current_scan_path}\n"
            msg += f"📈 文件夹数量: {len(results)}\n"
            msg += f"💾 总大小: {self._format_size(total_size)}\n"
            msg += f"🏆 最大文件夹: {largest_name} ({largest})"
            
            QMessageBox.information(self, "扫描完成", msg)
    
    def scan_error(self, error_msg):
        """扫描错误"""
        QMessageBox.critical(self, "扫描错误", f"❌ 扫描过程中发生错误:\n{error_msg}")
        self.scan_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.export_button.setEnabled(False)  # 扫描错误时禁用导出按钮
        self.statusBar().showMessage("❌ 扫描失败")
    
    def stop_scan(self):
        """停止扫描"""
        if self.scanner_thread and self.scanner_thread.isRunning():
            self.scanner_thread.cancel()
        self.scanner_thread.wait()
        self.scan_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.export_button.setEnabled(False)  # 扫描停止时禁用导出按钮
        self.statusBar().showMessage("⏹️ 扫描已停止")
    
    def export_to_excel(self):
        """将扫描结果导出到Excel文件"""
        # 检查是否有扫描结果
        if not self.table_model.items:
            QMessageBox.warning(self, "导出失败", "没有可导出的数据，请先执行扫描")
            return
        
        # 尝试导入Excel库
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            # 如果openpyxl不可用，尝试使用xlsxwriter
            try:
                import xlsxwriter
            except ImportError:
                QMessageBox.critical(self, "导出失败", "无法导出到Excel，请先安装openpyxl或xlsxwriter库")
                return
        
        # 获取保存路径
        from PySide6.QtWidgets import QFileDialog
        file_path, _ = QFileDialog.getSaveFileName(self, "导出到Excel", "扫描结果.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return  # 用户取消了保存
        
        try:
            # 尝试使用openpyxl导出
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "扫描结果"
                
                # 设置表头
                headers = ['序号', '名称', '类型', '路径', '大小', '百分比']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 填充数据
                for row, item in enumerate(self.table_model.items, 2):
                    ws.cell(row=row, column=1, value=row-1)
                    ws.cell(row=row, column=2, value=item['name'])
                    ws.cell(row=row, column=3, value='文件夹' if item['type'] == 'folder' else '文件')
                    ws.cell(row=row, column=4, value=item['path'])
                    ws.cell(row=row, column=5, value=item['display_size'])
                    ws.cell(row=row, column=6, value=item['size'] / (1024**3) if item['size'] > 0 else 0)
                
                # 调整列宽
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[chr(64 + col)].auto_size = True
                
                wb.save(file_path)
                wb.close()
            except ImportError:
                # 如果openpyxl失败，使用xlsxwriter
                workbook = xlsxwriter.Workbook(file_path)
                worksheet = workbook.add_worksheet("扫描结果")
                
                # 设置表头格式
                header_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter'})
                
                # 设置表头
                headers = ['序号', '名称', '类型', '路径', '大小', '百分比']
                worksheet.write_row(0, 0, headers, header_format)
                
                # 填充数据
                for row, item in enumerate(self.table_model.items, 1):
                    worksheet.write(row, 0, row)
                    worksheet.write(row, 1, item['name'])
                    worksheet.write(row, 2, '文件夹' if item['type'] == 'folder' else '文件')
                    worksheet.write(row, 3, item['path'])
                    worksheet.write(row, 4, item['display_size'])
                    worksheet.write(row, 5, item['size'] / (1024**3) if item['size'] > 0 else 0)
                
                # 调整列宽
                worksheet.set_column('A:A', 8)
                worksheet.set_column('B:B', 25)
                worksheet.set_column('C:C', 10)
                worksheet.set_column('D:D', 50)
                worksheet.set_column('E:E', 15)
                worksheet.set_column('F:F', 15)
                
                workbook.close()
            
            QMessageBox.information(self, "导出成功", f"扫描结果已成功导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出过程中发生错误:\n{str(e)}")
    
    def open_selected_folder(self):
        """打开选中的文件夹（从树形视图）"""
        # 尝试从树形视图打开
        tree_index = self.tree_view.currentIndex()
        if tree_index.isValid():
            item = self.tree_model.itemFromIndex(tree_index)
            path = item.data(Qt.UserRole)
            if path and os.path.exists(path):
                self._open_explorer(path)
                return
        
        # 尝试从表格视图打开
        table_index = self.table_view.currentIndex()
        if table_index.isValid():
            self.open_folder_from_table(table_index)
            return
        
        QMessageBox.warning(self, "警告", "请先选择一个文件夹")
    
    def open_folder_from_table(self, index):
        """从表格打开文件夹"""
        source_index = self.table_proxy.mapToSource(index)
        if source_index.isValid():
            folder = self.table_model.items[source_index.row()]
            path = folder.get('path', '')
            if path and os.path.exists(path):
                self._open_explorer(path)
    
    def delete_selected_items(self):
        """删除选中的文件或文件夹到回收站"""
        # 获取选中的行
        selected_rows = self.table_view.selectionModel().selectedRows()
        if not selected_rows:
            return
        
        # 确认删除操作
        confirm = QMessageBox.question(self, "确认删除", 
                                      f"确定要将选中的 {len(selected_rows)} 个项目删除到回收站吗？\n\n注意：文件夹将被彻底删除，文件将被删除到回收站。",
                                      QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if confirm != QMessageBox.Yes:
            return
        
        # 准备Windows API函数
        try:
            # Windows API常量
            FO_DELETE = 3
            FOF_ALLOWUNDO = 0x40  # 删除到回收站
            FOF_NOCONFIRMATION = 0x10  # 不显示确认对话框
            FOF_SILENT = 0x4  # 不显示进度对话框
            
            # 定义结构体
            class SHFILEOPSTRUCT(ctypes.Structure):
                _fields_ = [
                    ("hwnd", wintypes.HWND),
                    ("wFunc", wintypes.UINT),
                    ("pFrom", ctypes.c_wchar_p),
                    ("pTo", ctypes.c_wchar_p),
                    ("fFlags", wintypes.WORD),
                    ("fAnyOperationsAborted", wintypes.BOOL),
                    ("hNameMappings", ctypes.c_void_p),
                    ("lpszProgressTitle", ctypes.c_wchar_p)
                ]
            
            # 获取函数
            shfileop = ctypes.windll.shell32.SHFileOperationW
            shfileop.argtypes = [ctypes.POINTER(SHFILEOPSTRUCT)]
            shfileop.restype = wintypes.INT
            
            # 处理选中的项目
            success_count = 0
            failed_count = 0
            failed_items = []
            
            for index in selected_rows:
                source_index = self.table_proxy.mapToSource(index)
                if source_index.isValid():
                    item = self.table_model.items[source_index.row()]
                    path = item.get('path', '')
                    
                    if not path or not os.path.exists(path):
                        failed_count += 1
                        failed_items.append(item.get('name', ''))
                        continue
                    
                    try:
                        # 构建SHFILEOPSTRUCT
                        file_op = SHFILEOPSTRUCT()
                        file_op.hwnd = None
                        file_op.wFunc = FO_DELETE
                        file_op.pFrom = path + '\0'  # 路径必须以双空字符结尾
                        file_op.pTo = None
                        file_op.fFlags = FOF_ALLOWUNDO | FOF_NOCONFIRMATION | FOF_SILENT
                        file_op.fAnyOperationsAborted = False
                        file_op.hNameMappings = None
                        file_op.lpszProgressTitle = None
                        
                        # 调用API
                        result = shfileop(ctypes.byref(file_op))
                        if result == 0 and not file_op.fAnyOperationsAborted:
                            success_count += 1
                        else:
                            failed_count += 1
                            failed_items.append(item.get('name', ''))
                    except Exception as e:
                        failed_count += 1
                        failed_items.append(item.get('name', ''))
            
            # 显示删除结果
            msg = f"删除完成！\n\n"
            msg += f"成功删除: {success_count} 个项目\n"
            if failed_count > 0:
                msg += f"删除失败: {failed_count} 个项目\n"
                if len(failed_items) <= 10:
                    msg += f"失败项目: {', '.join(failed_items)}"
                else:
                    msg += f"失败项目: {', '.join(failed_items[:10])}... 等{failed_count}个"
            
            QMessageBox.information(self, "删除结果", msg)
            
        except Exception as e:
            QMessageBox.critical(self, "删除失败", f"删除过程中发生错误:\n{str(e)}")
    
    def _open_explorer(self, path):
        """使用系统资源管理器打开文件夹"""
        try:
            if os.name == 'nt':  # Windows
                # 使用explorer打开并选中
                if os.path.isfile(path):
                    subprocess.Popen(f'explorer /select,"{path}"')
                else:
                    subprocess.Popen(f'explorer "{path}"')
            else:  # macOS/Linux
                subprocess.Popen(['xdg-open', path])
        except Exception as e:
            QMessageBox.warning(self, "打开失败", f"无法打开文件夹:\n{str(e)}")
    
    def show_tree_context_menu(self, position):
        """显示树形视图的右键菜单"""
        index = self.tree_view.indexAt(position)
        if index.isValid():
            menu = QMenu()
            
            open_action = QAction("📂 打开文件夹", self)
            open_action.triggered.connect(self.open_selected_folder)
            menu.addAction(open_action)
            
            scan_action = QAction("🔍 扫描此文件夹", self)
            scan_action.triggered.connect(self.start_scan)
            menu.addAction(scan_action)
            
            menu.addSeparator()
            
            refresh_action = QAction("🔄 刷新", self)
            refresh_action.triggered.connect(self.refresh_tree_item)
            menu.addAction(refresh_action)
            
            expand_action = QAction("📖 展开所有子文件夹", self)
            expand_action.triggered.connect(lambda: self.expand_tree_item(index))
            menu.addAction(expand_action)
            
            menu.exec_(self.tree_view.viewport().mapToGlobal(position))
    
    def show_table_context_menu(self, position):
        """显示表格视图的右键菜单"""
        index = self.table_view.indexAt(position)
        if index.isValid():
            menu = QMenu()
            
            open_action = QAction("📂 打开文件夹/文件", self)
            open_action.triggered.connect(lambda: self.open_folder_from_table(index))
            menu.addAction(open_action)
            
            menu.addSeparator()
            
            copy_path_action = QAction("📋 复制路径", self)
            copy_path_action.triggered.connect(lambda: self.copy_path_from_table(index))
            menu.addAction(copy_path_action)
            
            copy_size_action = QAction("📊 复制大小", self)
            copy_size_action.triggered.connect(lambda: self.copy_size_from_table(index))
            menu.addAction(copy_size_action)
            
            menu.addSeparator()
            
            # 删除选中选项
            delete_action = QAction("🗑️ 删除选中", self)
            delete_action.triggered.connect(self.delete_selected_items)
            menu.addAction(delete_action)
            
            menu.addSeparator()
            
            locate_action = QAction("📍 在树形图中定位", self)
            locate_action.triggered.connect(lambda: self.locate_in_tree(index))
            menu.addAction(locate_action)
            
            menu.exec_(self.table_view.viewport().mapToGlobal(position))
    
    def copy_path_from_table(self, index):
        """复制路径到剪贴板"""
        source_index = self.table_proxy.mapToSource(index)
        if source_index.isValid():
            item = self.table_model.items[source_index.row()]
            path = item.get('path', '')
            if path:
                clipboard = QApplication.clipboard()
                clipboard.setText(path)
                self.statusBar().showMessage("路径已复制到剪贴板", 2000)
    
    def copy_size_from_table(self, index):
        """复制大小到剪贴板"""
        source_index = self.table_proxy.mapToSource(index)
        if source_index.isValid():
            item = self.table_model.items[source_index.row()]
            size = item.get('display_size', '')
            if size:
                clipboard = QApplication.clipboard()
                clipboard.setText(size)
                self.statusBar().showMessage("大小已复制到剪贴板", 2000)
    
    def refresh_tree_item(self):
        """刷新树节点"""
        current_index = self.tree_view.currentIndex()
        if current_index.isValid():
            item = self.tree_model.itemFromIndex(current_index)
            
            # 移除所有子项
            item.removeRows(0, item.rowCount())
            
            # 重新加载
            self.load_subfolders(item)
    
    def expand_tree_item(self, index):
        """展开树节点的所有子文件夹"""
        item = self.tree_model.itemFromIndex(index)
        
        # 先加载子文件夹
        if item.rowCount() == 1:
            child = item.child(0)
            if child and child.text() == "...":
                item.removeRow(0)
                self.load_subfolders(item)
        
        # 展开当前节点
        self.tree_view.expand(index)
        
        # 递归展开所有子节点
        for i in range(item.rowCount()):
            child_item = item.child(i)
            if child_item:
                self.expand_tree_item(child_item.index())
    
    def locate_in_tree(self, index):
        """在树形图中定位文件夹"""
        source_index = self.table_proxy.mapToSource(index)
        if source_index.isValid():
            folder = self.table_model.folders[source_index.row()]
            path = folder.get('path', '')
            
            # 在树形图中查找并选中该路径
            self.select_path_in_tree(path)
    
    def select_path_in_tree(self, path):
        """在树形图中选择指定路径"""
        # 遍历树模型查找路径
        for i in range(self.tree_model.rowCount()):
            disk_item = self.tree_model.item(i)
            if self._find_and_select_path(disk_item, path):
                return
    
    def _find_and_select_path(self, item, target_path):
        """递归查找并选择路径"""
        current_path = item.data(Qt.UserRole)
        if current_path and current_path.lower() == target_path.lower():
            # 展开父节点
            parent = item.parent()
            if parent:
                self.tree_view.expand(parent.index())
            
            # 选中当前项
            selection_model = self.tree_view.selectionModel()
            selection_model.select(item.index(), QItemSelectionModel.ClearAndSelect)
            self.tree_view.scrollTo(item.index())
            return True
        
        # 递归查找子项
        for i in range(item.rowCount()):
            child = item.child(i)
            if self._find_and_select_path(child, target_path):
                # 展开当前节点
                self.tree_view.expand(item.index())
                return True
        
        return False
    
    def refresh_disks(self):
        """刷新磁盘列表"""
        self.load_disks()
        if self.disk_combo.count() > 0:
            self.on_disk_changed(self.disk_combo.currentIndex())
    
    def _format_size(self, size_bytes):
        """格式化文件大小显示"""
        if size_bytes == 0:
            return "0 B"
        
        size_names = ("B", "KB", "MB", "GB", "TB")
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1
        
        if i == 0:
            return f"{int(size_bytes)} B"
        elif i == 1:
            return f"{size_bytes:.1f} KB"
        elif i == 2:
            return f"{size_bytes:.1f} MB"
        elif i == 3:
            return f"{size_bytes:.2f} GB"
        else:
            return f"{size_bytes:.2f} TB"
    
    def closeEvent(self, event):
        """窗口关闭事件"""
        if self.scanner_thread and self.scanner_thread.isRunning():
            self.scanner_thread.cancel()
            self.scanner_thread.wait()
        event.accept()

def main():
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create('Fusion'))
    
    # 设置深色调色板
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor(30, 30, 30))
    palette.setColor(QPalette.WindowText, QColor(220, 220, 220))
    palette.setColor(QPalette.Base, QColor(20, 20, 20))
    palette.setColor(QPalette.AlternateBase, QColor(40, 40, 40))
    palette.setColor(QPalette.ToolTipBase, QColor(40, 40, 40))
    palette.setColor(QPalette.ToolTipText, QColor(220, 220, 220))
    palette.setColor(QPalette.Text, QColor(220, 220, 220))
    palette.setColor(QPalette.Button, QColor(50, 50, 50))
    palette.setColor(QPalette.ButtonText, QColor(220, 220, 220))
    palette.setColor(QPalette.BrightText, QColor(255, 255, 255))
    palette.setColor(QPalette.Highlight, QColor(187, 134, 252))
    palette.setColor(QPalette.HighlightedText, QColor(0, 0, 0))
    app.setPalette(palette)
    
    # 设置应用程序信息
    app.setApplicationName("磁盘空间分析工具")
    app.setOrganizationName("DiskAnalyzer")
    
    window = DarkDiskSpaceAnalyzer()
    window.show()
    
    sys.exit(app.exec())

if __name__ == '__main__':
    main()